#!/usr/bin/env python3
# -*- coding: utf-8 -*-

#import sys
#import csv
#import matplotlib.pyplot as plt
import numpy as np
#import cv2
import xlsxwriter as xlw
from openpyxl import load_workbook
import sys
import matplotlib.pyplot as plt
import cv2

def bb_intersection_over_union(boxA, boxB):
    # determine the (x, y)-coordinates of the intersection rectangle
    xA = max(boxA[0], boxB[0])
    yA = max(boxA[1], boxB[1])
    xB = min(boxA[2], boxB[2])
    yB = min(boxA[3], boxB[3])
    
    # compute the area of intersection rectangle
    interArea = (xB - xA + 1) * (yB - yA + 1)
    
    # compute the area of both the prediction and ground-truth
    # rectangles
    boxAArea = (boxA[2] - boxA[0] + 1) * (boxA[3] - boxA[1] + 1)
    boxBArea = (boxB[2] - boxB[0] + 1) * (boxB[3] - boxB[1] + 1)
    
    # compute the intersection over union by taking the intersection
    # area and dividing it by the sum of prediction + ground-truth
    # areas - the interesection area
    iou = max(0,interArea / float(boxAArea + boxBArea - interArea))
    if iou>1.0:    
        iou = 0.0
    
    # return the intersection over union value
    return iou

imgFolder = "/Users/Greenborg/Desktop/Experiments/Test/"
file = "/Users/Greenborg/Desktop/Experiments/Test/"
fileEval = "/Users/Greenborg/Desktop/Source/Nuclei-CNTK/FasterRCNN/Output/RefNew/Results/All/"
name = "RefNoNMS.txt"
SaveN = "RefNoNMS"
#num = 5;

#Floc = open(file+"testval_nuclei.txt","r")
#Froi = open(file+"testval_nuclei_roi.txt","r")
#
#lines = Floc.readlines()
#roiLines = Froi.readlines()
#
#Floc.close()
#Froi.close()
#
#
#FroiEval = open(fileEval+name)
#roiLinesEval = FroiEval.readlines()
#FroiEval.close()

wbIntern = load_workbook(fileEval+'RefNoNMS_InternDist.xlsx',read_only=True)
wbInternGT = load_workbook(fileEval+'RefNoNMS_InternGT.xlsx',read_only=True)



#    PrecTop = np.zeros([4000,100])
#    RecTop = np.zeros([4000,100])
#    PrecTopP = np.zeros([4000,100])
#    RecTopP = np.zeros([4000,100])
#    PrecTopN = np.zeros([4000,100])
#    RecTopN = np.zeros([4000,100])
Prec = []
Rec = []
PrecN = []
RecN = []
PrecP = []
RecP = []
PosNegSum = np.zeros(2)
PosNegGT = np.zeros(2)
PosNegPred = np.zeros(2)
iNum = 1341#int(sys.argv[1])
    #lineLoc = lines[i]
    #loc = lineLoc.split()
    
#        lineRoi = roiLines[i]
#        roi = lineRoi.split()
#        roi[0:2] = []
#        
#        lineRoiEval = roiLinesEval[i]
#        roiEval = lineRoiEval.split()
#        
#        distArrayIntern = np.zeros([int(len(roiEval)/5),int(len(roiEval)/5)])
#        distArray = np.zeros([int(len(roi)/5),int(len(roiEval)/6)])
    #roiSel = np.zeros(int(len(roi)))
    #roiEvalSel = np.zeros(int(len(roiEval)))
for i in range(iNum,iNum+1):
    rows = np.zeros([wbIntern[str(i)].max_row,wbIntern[str(i)].max_column])
    
    ws = wbIntern[str(i)]
    k=0
    for row in ws.rows:
        j=0
        for cell in row:
            rows[k,j]=cell.value
            j+=1
        k+=1
    
    rowsGTRaw = np.zeros([wbInternGT[str(i)].max_row,wbInternGT[str(i)].max_column])
    ws = wbInternGT[str(i)]
    k=0
    for row in ws.rows:
        j=0
        for cell in row:
            rowsGTRaw[k,j]=cell.value
            j+=1
        k+=1
    
    stopper = True
    while stopper:
        if not rows[-1][0]:
            del(rows[-1])
        else:
            stopper=False
    stopper = True
    #        while stopper:
    #            if not rowsGT[-1][0]:
    #                del(rowsGT[-1])
    #            else:
    #                stopper=False        
    
    
    score = [rows[j][0] for j in range(wbIntern[str(i)].max_row)]
    PosNeg = [rows[j][1] for j in range(wbIntern[str(i)].max_row)]
    PosNeg = np.asarray(PosNeg)
    sorter = np.argsort(score)
    sorter = sorter[::-1]
    NumGT = np.size(rowsGTRaw,0)
    NumPro = np.size(rows,0)
    NMS = 0.35
    NMS2 = 0.35
    thr = 0.74
    thr2 = 0.9
    NMSTP = 0.5
    NMSTP2 = 1.0
                
    keepPred = np.zeros(NumPro)
    notKeepPred = np.zeros(NumPro)
    removeThres = np.zeros(NumPro)
    
    for k in range(NumPro):
        ind = sorter[k]
        if notKeepPred[ind]==0:
            if score[ind]>thr:
                keepPred[ind] = 1
                for j in range(NumPro):
                    if rows[ind][j+2]>NMS:
                        if PosNeg[ind]==PosNeg[j]:
                            if keepPred[j]==0:
                                notKeepPred[j]=1
                                #if score[ind]<thr:
                                    #removeThres[j] = 1
            elif score[ind]>thr2:
                removeThres[j] = 1
    
    for w in range(1):
        rowsGT = rowsGTRaw.copy()
        truePred = np.zeros(NumPro)
        GTTrue = []
        for k in range(NumGT):
            nexter = True
            while nexter:
                maxer = np.argmax(rowsGT[k][1:np.size(rowsGT,1)])
                if rowsGT[k][maxer+1]>NMSTP and rowsGT[k][maxer+1]<=NMSTP2:
                    if keepPred[maxer]==1:
                        if PosNeg[maxer]==rowsGT[k][0]:
                            if truePred[maxer]==0:
                                truePred[maxer]=1
                                nexter=False
                                GTTrue.append(k)
                            else:
                                rowsGT[k][maxer+1]=0.0
                        else:
                            rowsGT[k][maxer+1]=0.0
                    else:
                        rowsGT[k][maxer+1]=0.0
                else:
                    nexter=False
                        
        if np.sum(keepPred)>0.0:
            Prec = np.sum(truePred)/np.sum(keepPred)
        else:
            Prec = 0
            
        Rec = np.sum(truePred)/NumGT
        
        PNPred = PosNeg[list(map(bool,keepPred))]
        inder = list(map(bool,PosNeg))
        inder2 = list(map(bool,abs(PosNeg-1)))
        
        PNGT = np.asarray([rowsGT[j][0] for j in range(wbInternGT[str(i)].max_row)] )
        
        PosNegPred[0] += np.sum(PNPred)
        PosNegPred[1]+= np.sum(abs(PNPred-1))
        
        PosNegSum[0] += np.sum(truePred[inder])
        PosNegSum[1] += np.sum(truePred[inder2])
        
        PosNegGT[0] += np.sum(PNGT)
        PosNegGT[1]+= np.sum(abs(PNGT-1))
        
        if np.sum(abs(PNPred-1))>0:
            PrecN = np.sum(truePred[inder2])/np.sum(abs(PNPred-1))
        else:
            PrecN = 0
        
        if np.sum(abs(PNGT-1))>0:
            RecN = np.sum(truePred[inder2])/np.sum(abs(PNGT-1))
        else:
            RecN = 0
            
        if np.sum(PNPred)>0:
            PrecP = np.sum(truePred[inder])/np.sum(PNPred)
        else:
            PrecP = 0
            
        if np.sum(PNGT)>0:
            RecP= np.sum(truePred[inder])/np.sum(PNGT)
        else:
            RecP= 0
        
        imgFolder = "/Users/Greenborg/Desktop/Experiments/Test/"
        #file = "/Users/Greenborg/Desktop/Experiments/"
        file = "/Users/Greenborg/Desktop/Experiments/Test/"
        file2 = "/Users/Greenborg/Desktop/Experiments/"
        fileEval = "/Users/Greenborg/Desktop/Source/Nuclei-CNTK/FasterRCNN/Output/"
        name = "/RefNew/RefNoNMS.txt"
        pltNum = 2
        
        #imgVal = num % 100
        #folderVal = (num - imgVal) / 100
        #imgVal = imgVal - 1
        
        Floc = open(file+"testval_nuclei.txt","r")
        Froi = open(file+"testval_nuclei_roi.txt","r")
        #FroiVis = open(file2+"testvis_nuclei_minmax_roi.txt","r")
        FroiVis = open(file2+"testvis_nuclei_mean25_roi.txt","r")
        
        
        lines = Floc.readlines()
        lineLoc = lines[i]
        loc = lineLoc.split()
        
        img = cv2.imread(imgFolder+loc[1],1)
        
        roiLines = Froi.readlines()
        lineRoi = roiLines[i]
        roi = lineRoi.split()
        roi[0:2] = []
        
        roiLinesVis = FroiVis.readlines()
        lineRoiVis = roiLinesVis[i]
        roiVis = lineRoiVis.split()
        roiVis[0:2] = []
        
        FroiEval = open(fileEval+name)
        roiLinesEval = FroiEval.readlines()
        lineRoiEval = roiLinesEval[i]
        roiEval = lineRoiEval.split()
        
        #numPlt = np.argwhere(keepPred==1)
        numPlt = np.argwhere(keepPred==1)
        for l in range(NumGT):#GTTrue:#range(int(len(roi)/5)):
            
            ind = int(l*5)
            if pltNum==1:
                cv2.rectangle(img,(round(float(roi[ind])),round(float(roi[ind+1]))),(round(float(roi[ind+2])),round(float(roi[ind+3]))),(0,255,0),1)
            elif pltNum ==2:
                cv2.circle(img,(int(round(np.mean((float(roi[ind]),float(roi[ind+2]))))),int(round(np.mean((float(roi[ind+1]),float(roi[ind+3])))))),3,(0,255,0),-1)
            elif pltNum ==3:
                cv2.rectangle(img,(round(float(roi[ind])),round(float(roi[ind+1]))),(round(float(roi[ind+2])),round(float(roi[ind+3]))),(0,255,0),1)
                cv2.circle(img,(int(round(np.mean((float(roi[ind]),float(roi[ind+2]))))),int(round(np.mean((float(roi[ind+1]),float(roi[ind+3])))))),3,(0,255,0),-1)
            else:
                if int(round(float(roi[ind+4]))) == 1:
                    cv2.circle(img,(int(round(np.mean((float(roi[ind]),float(roi[ind+2]))))),int(round(np.mean((float(roi[ind+1]),float(roi[ind+3])))))),3,(0,255,0),-1)
                else:
                    cv2.circle(img,(int(round(np.mean((float(roi[ind]),float(roi[ind+2]))))),int(round(np.mean((float(roi[ind+1]),float(roi[ind+3])))))),3,(0,0,255),-1)
        
#        
        for l in range(int(np.sum(keepPred))):
            ind = int(numPlt[l]*6)
            if pltNum==1:
                cv2.rectangle(img,(round(float(roiEval[ind])),round(float(roiEval[ind+1]))),(round(float(roiEval[ind+2])),round(float(roiEval[ind+3]))),(0,0,255),1)
            elif pltNum==2:
                cv2.circle(img,(int(round(np.mean((float(roiEval[ind]),float(roiEval[ind+2]))))),int(round(np.mean((float(roiEval[ind+1]),float(roiEval[ind+3])))))),2,(0,0,255),-1)
            elif pltNum==3:
                cv2.rectangle(img,(round(float(roiEval[ind])),round(float(roiEval[ind+1]))),(round(float(roiEval[ind+2])),round(float(roiEval[ind+3]))),(0,0,255),1)
                cv2.circle(img,(int(round(np.mean((float(roiEval[ind]),float(roiEval[ind+2]))))),int(round(np.mean((float(roiEval[ind+1]),float(roiEval[ind+3])))))),2,(0,0,255),-1)
            else:
                if int(round(float(roiEval[ind+4]))) == 1:
                    cv2.circle(img,(int(round(np.mean((float(roiEval[ind]),float(roiEval[ind+2]))))),int(round(np.mean((float(roiEval[ind+1]),float(roiEval[ind+3])))))),3,(0,255,0),-1)
                else:
                    cv2.circle(img,(int(round(np.mean((float(roiEval[ind]),float(roiEval[ind+2]))))),int(round(np.mean((float(roiEval[ind+1]),float(roiEval[ind+3])))))),3,(0,0,255),-1)

#        numPlt = np.argwhere(removeThres==1)
#        for l in range(int(np.sum(removeThres))):
#            ind = int(numPlt[l]*6)
#            if pltNum==1:
#                cv2.rectangle(img,(round(float(roiEval[ind])),round(float(roiEval[ind+1]))),(round(float(roiEval[ind+2])),round(float(roiEval[ind+3]))),(255,0,0),1)
#            elif pltNum==2:
#                cv2.circle(img,(int(round(np.mean((float(roiEval[ind]),float(roiEval[ind+2]))))),int(round(np.mean((float(roiEval[ind+1]),float(roiEval[ind+3])))))),2,(255,0,0),-1)
#            elif pltNum==3:
#                cv2.rectangle(img,(round(float(roiEval[ind])),round(float(roiEval[ind+1]))),(round(float(roiEval[ind+2])),round(float(roiEval[ind+3]))),(255,0,0),1)
#                cv2.circle(img,(int(round(np.mean((float(roiEval[ind]),float(roiEval[ind+2]))))),int(round(np.mean((float(roiEval[ind+1]),float(roiEval[ind+3])))))),2,(255,0,0),-1)
#            else:
#                if int(round(float(roiEval[ind+4]))) == 1:
#                    cv2.circle(img,(int(round(np.mean((float(roiEval[ind]),float(roiEval[ind+2]))))),int(round(np.mean((float(roiEval[ind+1]),float(roiEval[ind+3])))))),3,(0,255,0),-1)
#                else:
#                    cv2.circle(img,(int(round(np.mean((float(roiEval[ind]),float(roiEval[ind+2]))))),int(round(np.mean((float(roiEval[ind+1]),float(roiEval[ind+3])))))),3,(0,0,255),-1)
#        
        for l in range(int(len(roi)/5)):#range(int(len(roi)/5)):
            
            ind = int(l*5)
            if pltNum==1:
                cv2.rectangle(img,(round(float(roiVis[ind])),round(float(roiVis[ind+1]))),(round(float(roiVis[ind+2])),round(float(roiVis[ind+3]))),(255,0,0),1)
            elif pltNum ==2:
                cv2.circle(img,(int(round(np.mean((float(roiVis[ind]),float(roiVis[ind+2]))))),int(round(np.mean((float(roiVis[ind+1]),float(roiVis[ind+3])))))),3,(255,0,0),-1)
            elif pltNum ==3:
                cv2.rectangle(img,(round(float(roiVis[ind])),round(float(roiVis[ind+1]))),(round(float(roiVis[ind+2])),round(float(roiVis[ind+3]))),(255,0,0),1)
                cv2.circle(img,(int(round(np.mean((float(roiVis[ind]),float(roiVis[ind+2]))))),int(round(np.mean((float(roiVis[ind+1]),float(roiVis[ind+3])))))),3,(255,0,0),-1)
            else:
                if int(round(float(roi[ind+4]))) == 1:
                    cv2.circle(img,(int(round(np.mean((float(roiVis[ind]),float(roiVis[ind+2]))))),int(round(np.mean((float(roiVis[ind+1]),float(roiVis[ind+3])))))),3,(0,255,0),-1)
                else:
                    cv2.circle(img,(int(round(np.mean((float(roiVis[ind]),float(roiVis[ind+2]))))),int(round(np.mean((float(roiVis[ind+1]),float(roiVis[ind+3])))))),3,(0,0,255),-1)
        
        if int(np.sum(truePred))>0:
            plt.imshow(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
            plt.show()
            #plt.savefig(fileEval+'RefNew/Results/All/Images/'+str(NMSTP2)+'/Dot'+str(i)+'.png')
       
        NMSTP -= 0.05
        NMSTP = round(NMSTP,2)
        NMSTP2 -= 0.05
        NMSTP2 = round(NMSTP2,2)
        
                