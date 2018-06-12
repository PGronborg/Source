#!/usr/bin/env python3
# -*- coding: utf-8 -*-

#import sys
#import csv
#import matplotlib.pyplot as plt
import numpy as np
#import cv2
import xlsxwriter as xlw
from openpyxl import load_workbook

def bb_intersection_over_union(boxA, boxB):
    # determine the (x, y)-coordinates of the intersection rectangle
    xA = max(boxA[0], boxB[0])
    yA = max(boxA[1], boxB[1])
    xB = min(boxA[2], boxB[2])
    yB = min(boxA[3], boxB[3])
    
    # compute the area of intersection rectangle
    interArea = (xB - xA + 1) * (yB - yA + 1)
    
    # compute the area of both the prediction and ground-truth
    # rectangles
    boxAArea = (boxA[2] - boxA[0] + 1) * (boxA[3] - boxA[1] + 1)
    boxBArea = (boxB[2] - boxB[0] + 1) * (boxB[3] - boxB[1] + 1)
    
    # compute the intersection over union by taking the intersection
    # area and dividing it by the sum of prediction + ground-truth
    # areas - the interesection area
    iou = max(0,interArea / float(boxAArea + boxBArea - interArea))
    if iou>1.0:    
        iou = 0.0
    
    # return the intersection over union value
    return iou

imgFolder = "/Users/Greenborg/Desktop/Experiments/Test/"
file = "/Users/Greenborg/Desktop/Experiments/Test/"
fileEval = "/Users/Greenborg/Desktop/Source/Nuclei-CNTK/FasterRCNN/Output/RefNew/Results/All/"
name = "RefNoNMS.txt"
SaveN = "RefNoNMS"
#num = 5;

#Floc = open(file+"testval_nuclei.txt","r")
#Froi = open(file+"testval_nuclei_roi.txt","r")
#
#lines = Floc.readlines()
#roiLines = Froi.readlines()
#
#Floc.close()
#Froi.close()
#
#
#FroiEval = open(fileEval+name)
#roiLinesEval = FroiEval.readlines()
#FroiEval.close()

wbIntern = load_workbook(fileEval+'RefNoNMS_InternDist.xlsx',read_only=True)
wbInternGT = load_workbook(fileEval+'RefNoNMS_InternGT.xlsx',read_only=True)



#    PrecTop = np.zeros([4000,100])
#    RecTop = np.zeros([4000,100])
#    PrecTopP = np.zeros([4000,100])
#    RecTopP = np.zeros([4000,100])
#    PrecTopN = np.zeros([4000,100])
#    RecTopN = np.zeros([4000,100])
Prec = np.zeros([4000,3,20])
Rec = np.zeros([4000,3,20])
PrecN = np.zeros([4000,3,20])
RecN = np.zeros([4000,3,20])
PrecP = np.zeros([4000,3,20])
RecP = np.zeros([4000,3,20])
PosNegSum = np.zeros([2,3,20])
PosNegGT = np.zeros([2,3,20])
PosNegPred = np.zeros([2,3,20])
for i in range(0,4000):
    print(i)
    #lineLoc = lines[i]
    #loc = lineLoc.split()
    
#        lineRoi = roiLines[i]
#        roi = lineRoi.split()
#        roi[0:2] = []
#        
#        lineRoiEval = roiLinesEval[i]
#        roiEval = lineRoiEval.split()
#        
#        distArrayIntern = np.zeros([int(len(roiEval)/5),int(len(roiEval)/5)])
#        distArray = np.zeros([int(len(roi)/5),int(len(roiEval)/6)])
    #roiSel = np.zeros(int(len(roi)))
    #roiEvalSel = np.zeros(int(len(roiEval)))
    
    rows = np.zeros([wbIntern[str(i)].max_row,wbIntern[str(i)].max_column])
    
    ws = wbIntern[str(i)]
    k=0
    for row in ws.rows:
        j=0
        for cell in row:
            rows[k,j]=cell.value
            j+=1
        k+=1
    
    rowsGTRaw = np.zeros([wbInternGT[str(i)].max_row,wbInternGT[str(i)].max_column])
    ws = wbInternGT[str(i)]
    k=0
    for row in ws.rows:
        j=0
        for cell in row:
            rowsGTRaw[k,j]=cell.value
            j+=1
        k+=1
    
    stopper = True
    while stopper:
        if not rows[-1][0]:
            del(rows[-1])
        else:
            stopper=False
    stopper = True
#        while stopper:
#            if not rowsGT[-1][0]:
#                del(rowsGT[-1])
#            else:
#                stopper=False        
    

    score = [rows[j][0] for j in range(wbIntern[str(i)].max_row)]
    PosNeg = [rows[j][1] for j in range(wbIntern[str(i)].max_row)]
    PosNeg = np.asarray(PosNeg)
    sorter = np.argsort(score)
    sorter = sorter[::-1]
    NumGT = np.size(rowsGTRaw,0)
    NumPro = np.size(rows,0)
    nn=1.0
    for o in range(3):
        nn=1.0
        index = 0
        while nn>0.04:
            rowsGT = rowsGTRaw.copy()
            if o==0:
                NMS = nn
                NMSTP = 0.5
                thr = 0.0
            elif o==1:
                NMS = 0.35
                NMSTP = nn
                thr = 0.74
            elif o==2:
                NMS = 0.35
                NMSTP = 0.5
                thr = nn-0.005
                
            keepPred = np.zeros(NumPro)
            notKeepPred = np.zeros(NumPro)
            
            for k in range(NumPro):
                ind = sorter[k]
                if notKeepPred[ind]==0:
                    if score[ind]>thr:
                        keepPred[ind] = 1
                        for j in range(NumPro):
                            if rows[ind][j+2]>NMS:
                                if PosNeg[ind]==PosNeg[j]:
                                    if keepPred[j]==0:
                                        notKeepPred[j]=1
            
            
            truePred = np.zeros(NumPro)
            for k in range(NumGT):
                nexter = True
                while nexter:
                    maxer = np.argmax(rowsGT[k][1:np.size(rowsGT,1)])
                    if rowsGT[k][maxer+1]>NMSTP:
                        if keepPred[maxer]==1:
                            if PosNeg[maxer]==rowsGT[k][0]:
                                if truePred[maxer]==0:
                                    truePred[maxer]=1
                                    nexter=False
                                else:
                                    rowsGT[k][maxer+1]=0.0
                            else:
                                rowsGT[k][maxer+1]=0.0
                        else:
                            rowsGT[k][maxer+1]=0.0
                    else:
                        nexter=False
                            
            if np.sum(keepPred)>0.0:
                Prec[i][o][index] = np.sum(truePred)/np.sum(keepPred)
            else:
                Prec[i][o][index] = 0
                
            Rec[i][o][index] = np.sum(truePred)/NumGT
            
            PNPred = PosNeg[list(map(bool,keepPred))]
            inder = list(map(bool,PosNeg))
            inder2 = list(map(bool,abs(PosNeg-1)))
            
            PNGT = np.asarray([rowsGT[j][0] for j in range(wbInternGT[str(i)].max_row)] )
            
            PosNegPred[0][o][index] += np.sum(PNPred)
            PosNegPred[1][o][index] += np.sum(abs(PNPred-1))
            
            PosNegSum[0][o][index] += np.sum(truePred[inder])
            PosNegSum[1][o][index] += np.sum(truePred[inder2])
            
            PosNegGT[0][o][index] += np.sum(PNGT)
            PosNegGT[1][o][index] += np.sum(abs(PNGT-1))
            
            if np.sum(abs(PNPred-1))>0:
                PrecN[i][o][index] = np.sum(truePred[inder2])/np.sum(abs(PNPred-1))
            else:
                PrecN[i][o][index] = 0
            
            if np.sum(abs(PNGT-1))>0:
                RecN[i][o][index] = np.sum(truePred[inder2])/np.sum(abs(PNGT-1))
            else:
                RecN[i][o][index] = 0
                
            if np.sum(PNPred)>0:
                PrecP[i][o][index] = np.sum(truePred[inder])/np.sum(PNPred)
            else:
                PrecP[i][o][index] = 0
                
            if np.sum(PNGT)>0:
                RecP[i][o][index] = np.sum(truePred[inder])/np.sum(PNGT)
            else:
                RecP[i][o][index] = 0
            nn-=0.05
            nn=round(nn,2)
            index +=1
            index = int(index)
                
#        k = 0
#        sorter = np.argsort(score)
#        sorter = sorter[::-1]
#        
#        Poser = Pro[pos.astype('bool')]
#        Nega = Pro[neg.astype('bool')]
#        sorterPos = np.argsort(score[pos.astype('bool')])
#        sorterNeg = np.argsort(score[neg.astype('bool')])
#        
#        PrecN[i] = np.sum(Nega)/np.sum(neg)
#        PrecP[i] = np.sum(Poser)/np.sum(pos)
#        
#        RecN[i] = np.sum(Nega)/np.sum(negGT)
#        RecP[i] = np.sum(Poser)/np.sum(posGT)
#        
#        PosNegGT[0] += np.sum(posGT)
#        PosNegGT[1] += np.sum(negGT)
#        PosNegPred[0] += np.sum(pos)
#        PosNegPred[1] += np.sum(neg)
#        PosNegSum[0] += np.sum(Poser)
#        PosNegSum[1] += np.sum(Nega)
#        
#        while k<100:
#            try:
#                PrecTop[i,k] = np.sum(Pro[sorter[0:(k+1)]])/(k+1)
#                RecTop[i,k] = np.sum(Pro[sorter[0:(k+1)]])/NumGT
#                k+=1
#            except:
#                break
#        
#        while k<100:
#            try:
#                PrecTopP[i,k] = np.sum(Poser[sorterPos[0:(k+1)]])/(k+1)
#                RecTopP[i,k] = np.sum(Poser[sorterPos[0:(k+1)]])/np.sum(posGT)
#                k+=1
#            except:
#                break
#        
#        while k<100:
#            try:
#                PrecTopN[i,k] = np.sum(Nega[sorterNeg[0:(k+1)]])/(k+1)
#                RecTopN[i,k] = np.sum(Nega[sorterNeg[0:(k+1)]])/np.sum(negGT)
#                k+=1
#            except:
#                break
nn=1.05        
for j in range(20):
    nn-=0.05
    nn=round(nn,2)
    for o in range(3):
            if o==0:
                namer = SaveN+'_NMS_'+str(round(nn*100))
            elif o==1:
                namer = SaveN+'_NMSTP_'+str(round(nn*100))
            elif o==2:
                namer = SaveN+'_Thres_'+str(round(nn*100))
                
            workbook = xlw.Workbook(fileEval+namer+'.xlsx')
            workS0 = workbook.add_worksheet('Facts')
            workS1 = workbook.add_worksheet('Pre Rec')
            workS2 = workbook.add_worksheet('Pos Pre Rec')
            workS3 = workbook.add_worksheet('Neg Pre Rec')
            
            workS0.write(0,1,'Positive')
            workS0.write(0,2,'Negative')
            workS0.write(0,3,'Total')
            
            workS0.write(1,0,'GT')
            workS0.write(1,1,PosNegGT[0][o][j])
            workS0.write(1,2,PosNegGT[1][o][j])
            workS0.write(1,3,PosNegGT[0][o][j]+PosNegGT[1][o][j])
            
            workS0.write(2,0,'Pred')
            workS0.write(2,1,PosNegPred[0][o][j])
            workS0.write(2,2,PosNegPred[1][o][j])
            workS0.write(2,3,PosNegPred[0][o][j]+PosNegPred[1][o][j])
            
            workS0.write(3,0,'Cor. pred')
            workS0.write(3,1,PosNegSum[0][o][j])
            workS0.write(3,2,PosNegSum[1][o][j])
            workS0.write(3,3,PosNegSum[0][o][j]+PosNegSum[1][o][j])
            
            workS1.write(0,1,'Precesion')
            workS1.write(0,2,'Recall')
            for i in range(0,4000):
                workS1.write(i+1, 0, '# '+str(i))
                workS1.write(i+1, 1, Prec[i][o][j])
                workS1.write(i+1, 2, Rec[i][o][j])
                workS2.write(i+1, 0, '# '+str(i))
                workS2.write(i+1, 1, PrecP[i][o][j])
                workS2.write(i+1, 2, RecP[i][o][j])
                workS3.write(i+1, 0, '# '+str(i))
                workS3.write(i+1, 1, PrecN[i][o][j])
                workS3.write(i+1, 2, RecN[i][o][j])
            workbook.close()

#while nn>0.01:
#    print(nn)
#    namer = SaveN+'_'+str(nn*100)
#    PrecRec(roiLines,lines,roiLinesEval,fileEval,namer,nn,0.5)
#    nn-=0.05
#
#nn=1.0
#while nn>0.01:
#    print(nn)
#    namer = SaveN+'TP_'+str(nn*100)
#    PrecRec(roiLines,lines,roiLinesEval,fileEval,namer,0.5,nn)
#    nn-=0.05    
    
    